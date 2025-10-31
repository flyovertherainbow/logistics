import streamlit as st
import pandas as pd
import numpy as np
import io
import re
from datetime import datetime
from openpyxl import load_workbook

# --- Constants and Utility Functions ---

# Mapping of keywords to standard column names (Case-insensitive matching)
COLUMN_MAP_A = {
    'Order #': 'Order #',
    'Supplier': 'Supplier',
    'Arrival Vessel': 'Arrival Vessel',
    'Arrival Voyage': 'Arrival Voyage',
    'ETA': 'ETA',
    'Container': 'Container'
}

COLUMN_MAP_B = {
    'BC PO': 'BC PO', # Covers BC PO/LC
    'Estimated Arrival': 'ETA', # Covers ETA Dates
    'Vessel': 'Arrival Vessel', # Covers Arrival Vessel, Vessel Name
    'Voyage': 'Arrival Voyage', # Covers Arrival Voyage, Voyage
    'Supplier': 'Supplier',
}

# Regex for container number: 4 upper letters followed by 7 digits
CONTAINER_NUMBER_PATTERN = re.compile(r'([A-Z]{4}\d{7})', re.IGNORECASE)
# Regex for container type: e.g., (20GP)
CONTAINER_TYPE_PATTERN = re.compile(r'\((20GP|20RE|40GP|40HC|40RE|40REHC)\)', re.IGNORECASE)


@st.cache_data
def detect_header_row(uploaded_file):
    """
    Finds the header row index (0-based) containing 'Order #' and 'Supplier' 
    in the first 20 rows of the TRI-STAR report (Excel A).
    """
    # Read entire file without a header first
    df_raw = pd.read_excel(uploaded_file, header=None, engine='openpyxl')
    
    keywords = ["Order #", "Supplier"]
    
    # Check up to the first 20 rows
    for i in range(min(20, len(df_raw))):
        # Convert row to string, check for keywords
        # Use str.lower() for case-insensitive matching
        row_str = ' '.join(df_raw.iloc[i].astype(str).str.lower().fillna(''))
        if all(keyword.lower() in row_str for keyword in keywords):
            return i
            
    return 0 # Fallback to first row


def get_latest_sheet_name(uploaded_file):
    """
    Selects the sheet in IMPORT DOC (Excel B) with the most recent date (MM.YYYY) 
    or the last sheet as a fallback.
    """
    try:
        # Load workbook to get sheet names
        wb = load_workbook(uploaded_file)
        sheet_names = wb.sheetnames
    except Exception:
        return None

    dated_sheets = {}
    for name in sheet_names:
        # Regex to find MM.YYYY format
        match = re.search(r'(\d{1,2}\.\d{4})', name)
        if match:
            try:
                # Try to parse as date (01 for day)
                date_obj = datetime.strptime(match.group(1), '%m.%Y')
                dated_sheets[name] = date_obj
            except ValueError:
                continue # Skip invalid dates

    if dated_sheets:
        # Return the sheet name corresponding to the latest date
        return max(dated_sheets, key=dated_sheets.get)
    
    # Fallback to the last sheet
    return sheet_names[-1] if sheet_names else None


def extract_pos(po_series: pd.Series) -> pd.Series:
    """
    Cleans and extracts 6-digit POs from a Series, handling combined and prefixed formats.
    Returns a Series where each element is a list of clean 6-digit PO strings.
    """
    def clean_po(po_str):
        if pd.isna(po_str) or po_str is None:
            return []
        
        po_str = str(po_str).upper().strip()
        
        # 1. Remove common prefixes like PO., PO#, PO
        po_str = re.sub(r'^(PO[#.]?)', '', po_str)
        
        # 2. Extract all 6-digit numbers. This handles:
        # - 107166
        # - 106815.A, 106815-1 (the non-digit part is ignored)
        # - 107070/107432 (both 6-digit numbers are captured)
        matches = re.findall(r'(\d{6})', po_str)
        
        clean_pos = set()
        for po in matches:
            if len(po) == 6:
                clean_pos.add(po)
                
        return sorted(list(clean_pos))

    # Apply the cleaning function and return as a Series
    return po_series.apply(clean_po)


def parse_container_string(container_str: str) -> tuple[str, str]:
    """Extracts the 4-letter + 7-digit container number and the container type."""
    if pd.isna(container_str) or not container_str:
        return None, None
        
    container_str = str(container_str).strip().upper()
    
    # Extract container number (4 letters + 7 digits)
    num_match = CONTAINER_NUMBER_PATTERN.search(container_str)
    container_number = num_match.group(1) if num_match else None
    
    # Extract container type (e.g., 20GP)
    type_match = CONTAINER_TYPE_PATTERN.search(container_str)
    container_type = type_match.group(1) if type_match else None
    
    return container_number, container_type

# --- Core Data Processing Functions ---

@st.cache_data
def process_excel_b(uploaded_file):
    """Processes IMPORT DOC (Excel B)"""
    with st.spinner("Processing IMPORT DOC..."):
        
        # 1. Sheet Selection
        sheet_name = get_latest_sheet_name(uploaded_file)
        if not sheet_name:
            st.error("No usable sheets found in IMPORT DOC.")
            return None

        st.info(f"Using latest sheet: **{sheet_name}**")
        
        df_b = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=0, engine='openpyxl')
        
        # Clean column names
        df_b.columns = [str(col).strip().replace('\n', ' ') for col in df_b.columns]
        b_cols_lower = {col.lower(): col for col in df_b.columns}
        
        # 2. Column Mapping and Standardization
        standard_cols = {}
        for keyword, std_name in COLUMN_MAP_B.items():
            found_col = next((b_col_name for lower_col, b_col_name in b_cols_lower.items() 
                              if keyword.lower() in lower_col), None)
            
            # Check for generic column names to prevent mapping errors on common words
            if found_col and std_name not in standard_cols.values():
                standard_cols[found_col] = std_name
                
        # Rename columns
        df_b = df_b.rename(columns=standard_cols)
        
        required_cols_b = ['BC PO', 'ETA']
        if not all(col in df_b.columns for col in required_cols_b):
            st.error(f"Missing essential columns in IMPORT DOC. Required: {required_cols_b}. Found: {df_b.columns.tolist()}")
            return None

        # 3. Container Consolidation
        # Find the first column containing 'container'
        container_start_col_name = next((col for col in df_b.columns if 'container' in col.lower()), None)
        cols_to_concat = []
        
        if container_start_col_name:
            col_index = df_b.columns.get_loc(container_start_col_name)
            
            # Start with the found container column
            cols_to_concat.append(container_start_col_name)
            
            # Check up to 5 subsequent columns
            for i in range(1, 6):
                if col_index + i < len(df_b.columns):
                    col_name = df_b.columns[col_index + i]
                    # Include if the name is an empty string or starts with 'Unnamed'
                    if not col_name or col_name.startswith('Unnamed'):
                        cols_to_concat.append(col_name)

        if cols_to_concat:
            # Fill NaN with empty string for concatenation, then join non-empty, stripped values with comma
            df_b['Container'] = df_b[cols_to_concat].astype(str).fillna('').agg(
                lambda x: ', '.join(filter(None, [v.strip() for v in x.tolist()])), axis=1
            )
        else:
            df_b['Container'] = '' # Create empty column if no container columns were found

        # PO extraction
        df_b['Clean_POs'] = extract_pos(df_b['BC PO'])
        df_b = df_b.explode('Clean_POs').rename(columns={'Clean_POs': 'PO'})
        df_b = df_b[df_b['PO'].notna() & (df_b['PO'].str.len() == 6)].copy()
        
        # Drop duplicates POs (keep the first entry for simplicity in mapping B data)
        df_b = df_b.drop_duplicates(subset=['PO'], keep='first')
        
        st.success(f"IMPORT DOC processed. Found **{len(df_b)}** unique POs.")
        return df_b.set_index('PO')


@st.cache_data
def process_excel_a(uploaded_file):
    """Processes TRI-STAR SHIPMENT REPORT (Excel A)"""
    with st.spinner("Processing TRI-STAR SHIPMENT REPORT..."):
        
        # 1. Header Detection
        header_row_index = detect_header_row(uploaded_file)
        
        # Read again with detected header row
        df_a = pd.read_excel(uploaded_file, header=header_row_index, engine='openpyxl')
        
        # Clean column names (strip leading/trailing space)
        df_a.columns = [str(col).strip() for col in df_a.columns]
        
        # Map to standard names
        a_cols_lower = {col.lower(): col for col in df_a.columns}
        standard_cols = {}
        
        for keyword, std_name in COLUMN_MAP_A.items():
            found_col = next((a_col_name for lower_col, a_col_name in a_cols_lower.items() 
                              if keyword.lower() in lower_col), None)
            if found_col:
                standard_cols[found_col] = std_name
                
        df_a = df_a.rename(columns=standard_cols)
        
        required_cols_a = list(COLUMN_MAP_A.values())
        if not all(col in df_a.columns for col in required_cols_a):
            st.error(f"Missing essential columns in TRI-STAR SHIPMENT REPORT. Required: {required_cols_a}. Found: {df_a.columns.tolist()}")
            return None

        # 2. Cleaning (Drop rows with invalid ETA)
        initial_rows = len(df_a)
        # Attempt to convert ETA to datetime, coercing errors to NaT
        df_a['ETA'] = pd.to_datetime(df_a['ETA'], errors='coerce')
        df_a = df_a.dropna(subset=['ETA']).copy()
        
        if len(df_a) < initial_rows:
            st.warning(f"Dropped {initial_rows - len(df_a)} rows with invalid 'ETA' in TRI-STAR SHIPMENT REPORT.")

        # PO extraction
        df_a['Clean_POs'] = extract_pos(df_a['Order #'])
        df_a = df_a.explode('Clean_POs').rename(columns={'Clean_POs': 'PO'})
        df_a = df_a[df_a['PO'].notna() & (df_a['PO'].str.len() == 6)].copy()

        # Drop duplicates POs (keep the first entry for simplicity in mapping A data)
        df_a = df_a.drop_duplicates(subset=['PO'], keep='first')
        
        st.success(f"TRI-STAR SHIPMENT REPORT processed. Found **{len(df_a)}** unique POs.")
        return df_a.set_index('PO')


@st.cache_data
def compare_dataframes(df_a: pd.DataFrame, df_b: pd.DataFrame):
    """Performs the core comparison logic."""
    
    # Discrepancy lists
    diff_eta = []
    diff_container = []
    
    # Identify matched POs
    matched_pos = sorted(list(set(df_a.index) & set(df_b.index)))
    
    for po in matched_pos:
        row_a = df_a.loc[po]
        row_b = df_b.loc[po]

        # --- ETA Comparison (Rule 3) ---
        eta_a = row_a['ETA'].normalize().date() if pd.notna(row_a['ETA']) else None
        
        # Convert Excel B's ETA to date, handling various inputs
        eta_b_raw = row_b.get('ETA')
        eta_b = None
        try:
            eta_b_dt = pd.to_datetime(eta_b_raw, errors='coerce')
            eta_b = eta_b_dt.normalize().date() if pd.notna(eta_b_dt) else None
        except:
            eta_b = None

        if eta_a != eta_b:
            diff_eta.append({
                'PO': po,
                'Vessel A (TRI-STAR)': str(row_a['Arrival Vessel']).strip(),
                'ETA A (TRI-STAR)': eta_a,
                'Vessel B (IMPORT DOC)': str(row_b.get('Arrival Vessel', '')).strip(),
                'ETA B (IMPORT DOC)': eta_b,
            })

        # --- Container Comparison (Rule 2) ---
        container_a_raw = row_a.get('Container')
        container_b_raw = row_b.get('Container')
        
        num_a, type_a = parse_container_string(container_a_raw)
        num_b, type_b = parse_container_string(container_b_raw)

        has_container_num_a = bool(num_a)
        has_container_num_b = bool(num_b)
        
        # Rule 2.3: If A has only container type, do nothing (num_a is None, but raw string contains type)
        if not has_container_num_a and type_a:
            continue
            
        # Rule 2.1: if A has container number, but B does not (B is empty/no number)
        if has_container_num_a and not has_container_num_b:
            diff_container.append({
                'PO': po,
                'Container (TRI-STAR)': f"{num_a} ({type_a})" if type_a else num_a,
                'Container (IMPORT DOC)': 'MISSING (No 4L7D number found)',
                'Reason': 'New Container Number (In TRI-STAR but missing in IMPORT DOC)'
            })
            continue

        # Rule 2.2: A and B have container numbers, but type is different
        if has_container_num_a and has_container_num_b:
            # Compare the extracted container numbers and types
            if num_a != num_b or type_a != type_b:
                 # Check if the discrepancy is only the container type
                reason = 'Different Container Details (Number or Type Mismatch)'
                if num_a == num_b and type_a != type_b:
                    reason = f'Container Type Mismatch: TRI-STAR has {type_a}, IMPORT DOC has {type_b}'
                    
                diff_container.append({
                    'PO': po,
                    'Container (TRI-STAR)': f"{num_a} ({type_a})" if type_a else num_a,
                    'Container (IMPORT DOC)': f"{num_b} ({type_b})" if type_b else num_b,
                    'Reason': reason
                })


    # 2. Unmatched POs (POs in A but not in B)
    po_a_only = sorted(list(set(df_a.index) - set(df_b.index)))
    unmatched_pos = []
    for po in po_a_only:
        row_a = df_a.loc[po]
        unmatched_pos.append({
            'PO': po,
            'Supplier': row_a['Supplier'],
            'ETA': row_a['ETA'].normalize().date() if pd.notna(row_a['ETA']) else None,
            'Arrival Vessel': row_a['Arrival Vessel'],
            'Container (TRI-STAR)': row_a['Container']
        })
    
    # Convert lists to DataFrames
    df_diff_eta = pd.DataFrame(diff_eta)
    df_diff_container = pd.DataFrame(diff_container)
    df_unmatched_pos = pd.DataFrame(unmatched_pos)

    return df_diff_eta, df_diff_container, df_unmatched_pos, len(matched_pos)


def convert_df_to_csv(df):
    """Converts DataFrame to a CSV string for download."""
    # Use io.StringIO to create an in-memory CSV buffer
    return df.to_csv(index=False).encode('utf-8')


def main():
    """Main Streamlit application function."""
    st.set_page_config(layout="wide")
    st.title("🚢 TRI-STAR SHIPMENT CHECK LIST")
    
    # File upload section
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### TRI-STAR SHIPMENT REPORT")
        file_a = st.file_uploader("Upload Excel A (Shipment Report)", type=['xlsx'], key='file_a')
        
    with col2:
        st.markdown("### IMPORT DOC")
        file_b = st.file_uploader("Upload Excel B (Import Doc)", type=['xlsx'], key='file_b')
        
    st.markdown("---")
    
    # Data Processing and Comparison
    if file_a and file_b:
        
        # Rerun button for cache clearing (useful during development/debugging)
        if st.button("Rerun Comparison (Clear Cache)", key='clear_cache', help="Click to force a fresh re-read and processing of the files."):
            st.cache_data.clear()
            st.experimental_rerun()
            return
            
        try:
            # Process files
            df_a = process_excel_a(file_a)
            df_b = process_excel_b(file_b)
            
            if df_a is None or df_b is None:
                return # Stop execution if processing failed

            # Compare DataFrames
            df_diff_eta, df_diff_container, df_unmatched_pos, matched_count = compare_dataframes(df_a, df_b)

            # --- Results Display and Download ---
            st.header("Comparison Discrepancies")
            st.markdown("---")

            # 1. Different in ETA
            st.subheader("1. 🗓️ Different in ETA")
            if not df_diff_eta.empty:
                st.error(f"**{len(df_diff_eta)}** POs have different ETAs.")
                null_highlight_color = '#f0f0f0'
                #st.dataframe(df_diff_eta.style.highlight_null(null_color='#f0f0f0'), use_container_width=True, hide_index=True)
                styled_df = df_diff_eta.style.applymap(
                    lambda val: f'background-color: {null_highlight_color}' if pd.isna(val) else ''
                )
            
                st.dataframe(styled_df, use_container_width=True, hide_index=True)

                # Downloadable CSV (Item 4)
                csv_eta = convert_df_to_csv(df_diff_eta)
                st.download_button(
                    label="Download Differences in ETA as CSV",
                    data=csv_eta,
                    file_name='eta_discrepancies.csv',
                    mime='text/csv',
                    key='download_eta'
                )
            else:
                st.success("✅ No ETA discrepancies found.")

            st.markdown("---")

            # 3. New Container Number (Container Discrepancy)
            st.subheader("2. 📦 Container Discrepancies")
            if not df_diff_container.empty:
                st.warning(f"**{len(df_diff_container)}** POs have container discrepancies (New Container or Type Mismatch).")
                st.dataframe(df_diff_container, use_container_width=True, hide_index=True)
                
                # Downloadable CSV (Item 6)
                csv_container = convert_df_to_csv(df_diff_container)
                st.download_button(
                    label="Download Container Discrepancies as CSV",
                    data=csv_container,
                    file_name='container_discrepancies.csv',
                    mime='text/csv',
                    key='download_container'
                )
            else:
                st.success("✅ No container discrepancies found.")

            st.markdown("---")

            # 2. Unmatched PO Number from TRI-STAR
            st.subheader("3. 📝 Unmatched PO Number (In TRI-STAR but not in IMPORT DOC)")
            if not df_unmatched_pos.empty:
                st.error(f"**{len(df_unmatched_pos)}** POs are in TRI-STAR but missing from IMPORT DOC.")
                st.dataframe(df_unmatched_pos.style.highlight_null(null_color='#f0f0f0'), use_container_width=True, hide_index=True)
                
                # Downloadable CSV (Item 5)
                csv_po = convert_df_to_csv(df_unmatched_pos)
                st.download_button(
                    label="Download Unmatched POs as CSV",
                    data=csv_po,
                    file_name='unmatched_pos.csv',
                    mime='text/csv',
                    key='download_po'
                )
            else:
                st.success("✅ All TRI-STAR POs were found in IMPORT DOC.")

            # Display match summary
            st.markdown("---")
            total_a = len(df_a)
            total_b = len(df_b)
            st.info(f"**Summary:** **{matched_count}** POs matched successfully out of {total_a} unique POs from TRI-STAR and {total_b} unique POs from IMPORT DOC.")

        except Exception as e:
            st.error("An unexpected error occurred during processing.")
            st.exception(e)
            
    elif file_a or file_b:
        st.warning("Please upload both Excel files to start the comparison.")
    else:
        st.info("Upload your TRI-STAR Shipment Report and Import Document files above to begin.")


if __name__ == '__main__':
    # Initialize main app function if this script is executed
    # Note: Streamlit environment will handle the execution of the main function
    main()
